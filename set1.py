import openpyxl
import openpyxl.styles as ops

file_xls = 'sravnenie ud.xlsx'
start_row = 2
data_col1 = 1
data_col2 = 2
sp1 = set()
sp2 = set()

wb = openpyxl.load_workbook(file_xls)
ws = wb.active
max_row = ws.max_row

for row in range(start_row, max_row+1):
    cell = ws.cell(row, data_col1)
    orig_content_cell = cell.value
    if orig_content_cell:
        content_cell = str(orig_content_cell.strip()).replace('.', '')
        if content_cell.isdigit():
            sp1.add(content_cell)
        else:
            cell.fill = ops.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    else:
        cell.fill = ops.PatternFill(start_color='000000', end_color='FF0000', fill_type='solid')

for row in range(start_row, max_row+1):
    cell = ws.cell(row, data_col2)
    orig_content_cell = cell.value
    if orig_content_cell:
        content_cell = str(orig_content_cell.strip()).replace('.', '')
        if content_cell.isdigit():
            sp2.add(content_cell)
        else:
            cell.fill = ops.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    else:
        cell.fill = ops.PatternFill(start_color='000000', end_color='FF0000', fill_type='solid')

for i in range(len(sp1)):
    ws.cell(start_row+i, data_col1+2).value = sp1
    pass



wb.save(file_xls)
