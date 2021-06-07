# сравнение номеров дел в двух столбцах и заполенение столбцов данными связанных множеств
# ...
# INSTALL
# pip install openpyxl
# COMPILE
# pyinstaller -F compare_ud.py

import openpyxl
import openpyxl.styles

# переменные для работы
file_xls = 'compare_ud.xlsx'
# строка с которой начинать собирать данные
start_row = 2
# колонки в которые пользователь вставляет данные
data_col1 = 1
data_col2 = 2
# два множества для данных из колонок со вставленными данными
set1 = set()
set2 = set()

# открывается существующая книга эксель
wb = openpyxl.load_workbook(file_xls)
ws = wb.active
# переменная самой последней заполненной строки
max_row = ws.max_row

# сбор, очистка и обработка данных из колонок с вставленными данными
# FF0000 красный - 000000 чёрный - 878787 серый
# колонка1
for row in range(start_row, max_row+1):
    cell = ws.cell(row, data_col1)
    content_cell = str(cell.value)
    content_cell = content_cell.strip().replace('.', '')
    content_cell = content_cell.strip().replace(' ', '')
    if content_cell.isdecimal():
        set1.add(content_cell)
    else:
        if len(content_cell) == 0:
            cell.fill = openpyxl.styles.PatternFill(start_color='878787', end_color='878787', fill_type='solid')
        elif cell.value is None:
            cell.fill = openpyxl.styles.PatternFill(start_color='878787', end_color='878787', fill_type='solid')
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# колонка2
for row in range(start_row, max_row+1):
    cell = ws.cell(row, data_col2)
    content_cell = str(cell.value)
    content_cell = content_cell.strip().replace('.', '')
    content_cell = content_cell.strip().replace(' ', '')
    if content_cell.isdecimal():
        set2.add(content_cell)
    else:
        if len(content_cell) == 0:
            cell.fill = openpyxl.styles.PatternFill(start_color='878787', end_color='878787', fill_type='solid')
        elif cell.value is None:
            cell.fill = openpyxl.styles.PatternFill(start_color='878787', end_color='878787', fill_type='solid')
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# заполнение колонок в экселе
# чистка третьей колонки - без дублей и пустых ячеек
list1 = list(set1)
list1.sort()
for list1_value in list1:
    ws.cell(start_row + list1.index(list1_value), 3).value = list1_value

# чистка четвёртой колонки - без дублей и пустых ячеек
list2 = list(set2)
list2.sort()
for list2_value in list2:
    ws.cell(start_row + list2.index(list2_value), 4).value = list2_value

# Объединение списков без дублей
all_cols = set1.union(set2)
list3 = list(all_cols)
list3.sort()
for list3_value in list3:
    ws.cell(start_row + list3.index(list3_value), 5).value = list3_value

# общие значения
common_values = set1.intersection(set2)
list4 = list(common_values)
list4.sort()
for list4_value in list4:
    ws.cell(start_row + list4.index(list4_value), 6).value = list4_value

# пересечение 1
diff_set1 = set1.difference(set2)
list5 = list(diff_set1)
list5.sort()
for list5_value in list5:
    ws.cell(start_row + list5.index(list5_value), 7).value = list5_value

# пересечение 2
diff_set2 = set2.difference(set1)
list6 = list(diff_set2)
list6.sort()
for list6_value in list6:
    ws.cell(start_row + list6.index(list6_value), 8).value = list6_value

# сохраняю исходный файл с заполненными ячейками
wb.save(file_xls)
